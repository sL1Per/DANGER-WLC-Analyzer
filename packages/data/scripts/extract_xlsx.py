#!/usr/bin/env python3
"""Extract reference data from the CLA/RPB xlsx exports into JSON."""
import json
import os
import re

import openpyxl

ROOT = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
OUT = os.path.join(os.path.dirname(__file__), "..", "json")
CLA = os.path.join(ROOT, "WoW Classic TBC - Combat Log Analytics V1.6.0a.xlsx")
RPB = os.path.join(ROOT, "WoW Classic TBC - Role Performance Breakdown V1.6.0a.xlsx")

# Fail fast if the source workbooks are missing.
for path in (CLA, RPB):
    if not os.path.exists(path):
        raise SystemExit(f"Missing source file: {path}")

def dump(name, obj):
    os.makedirs(OUT, exist_ok=True)
    # Sort dict outputs by numeric key so JSON diffs stay stable.
    if isinstance(obj, dict):
        obj = dict(sorted(obj.items(), key=lambda kv: int(kv[0])))
    with open(os.path.join(OUT, name), "w") as f:
        json.dump(obj, f, indent=1, ensure_ascii=False, sort_keys=False)
    print(f"wrote {name}: {len(obj)} entries")

def id_value_sheet(ws, skip_header):
    """Sheets shaped: col A = numeric id, col B = numeric value."""
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if skip_header and i == 0:
            continue
        a, b = row[0], row[1]
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            out[str(int(a))] = b if b != int(b) else int(b)
    return out

# Parse a cell that may contain a comma-separated list of NPC ids, including float-valued cells.
def parse_ids(cell):
    out = []
    for part in str(cell).split(","):
        part = part.strip()
        try:
            out.append(int(float(part)))
        except ValueError:
            pass
    return out

cla = openpyxl.load_workbook(CLA, data_only=True)
rpb = openpyxl.load_workbook(RPB, data_only=True)

# item id -> number of sockets ('sockets' has a header row)
dump("item-sockets.json", id_value_sheet(cla["sockets"], skip_header=True))
# item id -> shadow resistance ('shadow resistance config' has no header)
dump("item-shadow-res.json", id_value_sheet(cla["shadow resistance config"], skip_header=False))
# spell id -> spell haste value (no header)
dump("spell-haste.json", id_value_sheet(rpb["spell haste config"], skip_header=False))

# cheap/bad enchants: CLA 'gear issues' B5:C...
# Two formats: "927 [8]" (enchantId + slot) or bare int (slot-agnostic enchant id)
enchants = []
ws = cla["gear issues"]
for row in ws.iter_rows(min_row=5, min_col=2, max_col=3, values_only=True):
    bid, name = row
    if not isinstance(name, str):
        continue
    if isinstance(bid, str):
        m = re.match(r"^(\d+) \[(\d+)\]$", bid.strip())
        if m:
            enchants.append({"enchantId": int(m.group(1)),
                             "slot": int(m.group(2)), "name": name.strip()})
    elif isinstance(bid, (int, float)):  # slot-agnostic enchant stored as a bare number
        enchants.append({"enchantId": int(bid), "slot": None, "name": name.strip()})
dump("bad-enchants.json", enchants)

# excluded/fun items: CLA 'gear issues' E5:F...
excluded = []
for row in ws.iter_rows(min_row=5, min_col=5, max_col=6, values_only=True):
    iid, name = row
    if iid is not None and isinstance(name, str):
        try:
            excluded.append({"itemId": int(float(str(iid))), "name": name.strip()})
        except ValueError:
            pass
dump("excluded-items.json", excluded)

# speedrun trash requirements: CLA 'validate' rows A6+ (ids | zone | name | min kills)
reqs = []
for row in cla["validate"].iter_rows(min_row=6, max_col=4, values_only=True):
    ids, zone, name, minimum = row
    if ids is None or name is None or minimum is None:
        continue
    npc_ids = parse_ids(ids)
    reqs.append({"zone": str(zone), "name": str(name),
                 "npcIds": npc_ids, "minKills": int(float(str(minimum)))})
dump("trash-requirements.json", reqs)
